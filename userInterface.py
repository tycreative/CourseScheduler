from openpyxl import load_workbook
from time import localtime, strftime
import sys
import os
import re

import utilityFunctions as util
import objectsModule as objects
import semestersData as semesters
import coursesData as courses
import prereqsModule as prereqs
import excelModule as excel
import plannerModule as planner

def main():
    try:
        # Get prerequisites file
        prereqsFile = util.getFile("Please specify prerequisites file: ", [".csv", ".txt"])
        # Grab required courses and update course database using prereqs file
        courses, count = prereqs.findPrereqs(prereqsFile)
        if courses == []:
            raise Exception("An error loading the prerequisites has occurred.")
        print(f"Loaded {len(courses)} required courses...")
        print(f"Found {count} associated prerequisites...")

        if not util.getOption("Would you like to continue (y/n)?", "Y", "N"):
            raise Exception("User canceled execution.")

        # Get Excel file
        excelFile = util.getFile("Please specify Excel file (or leave blank): ", [".xlsx", ".xlsm", ".xltx", ".xltm"], True)
        # Determine if blank (and using template) or processing Excel file
        if excelFile != "":
            # Read from provided Excel file
            wb = load_workbook(excelFile)
            count = excel.findSemesters(wb.active)
            print(f"Found {len(semesters.getSemesters())} populated semesters...")
            print(f"Loaded {count} associated courses...")

            if not util.getOption("Would you like to continue (y/n)?", "Y", "N"):
                raise Exception("User canceled execution.")
            print() # just for consistent spacing lol

        print("1. Generate a course plan based on input files.")
        print("2. Check for prerequisite issues in existing course plan.")

        # User has selected option 1
        if util.getOption("Please select from the options above: ", "1", "2"):
            # Reset these in case user entered in files but wants to generate schedule instead
            semesters.semesters = {}
            wb = load_workbook("template.xlsx")
            # wb = load_workbook(os.path.join(sys._MEIPASS, "template.xlsx")) !!! USE THIS WHEN PACKAGING !!!
            

            # Sort courses by prereqs
            sortedPrereqs = prereqs.sortPrereqs([], courses)
            if sortedPrereqs == []:
                raise Exception("An error sorting the prerequisites has occurred.")
            print(f"Sorted {len(sortedPrereqs)} courses by prerequisites...")
            # Ask to include summer semesters
            summer = util.getOption("Would you like to include summer semesters (y/n)?", "Y", "N")
            year = util.getNumber("Please enter in a starting year:", 2000, 8000)
            maximum = util.getNumber("Please specify maximum amount of courses per semester:", 2, 7)
            print("\nCalculating...\n")

            # Generate semesters and their courses, then normalize
            planner.generateSemesters(sortedPrereqs, year)
            planner.normalizeSemesters(summer, maximum)
            # Write to output file
            output = "output_" + strftime("%H_%M_%S", localtime()) + ".xlsx"
            excel.outputExcel(output, wb, wb.active)
            print(f"Generated course plan has been outputted under '{output}'.")

        # User has selected option 2
        else:
            # Determine if provided excel input is empty
            if len(semesters.getSemesters()) == 0 or count == 0:
                raise Exception("Provided Excel input contains no semesters.")
            # Check for prereq issues
            print("\nChecking...\n")
            missing = prereqs.findMissing(courses.copy())
            issues = prereqs.findIssues()

            # Display any missing prerequisites
            if missing == []:
                print("No prerequisites missing...")
            else:
                print(f"{len(missing)} prerequisites missing:\n" + util.joinList(missing, "and"))
            # Display any prerequisite issues
            if issues == {}:
                print("\nNo prerequisite issues...")
            else:
                print(f"\n{len(issues)} prerequisite issues:")
                for course, issue in issues.items():
                    print(f"{course} missing: " + util.joinList(issue, "and"))

    # Display any exceptions raised in execution
    except Exception as e:
        print(e)
        return

    finally:
        input("Done.\nPress any key to exit.")

if __name__ == "__main__":
    main()