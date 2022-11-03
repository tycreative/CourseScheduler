from openpyxl.utils import get_column_letter
from coursesData import getCourse
import semestersData as data
import re



# Function to find semesters and their associated classes from an excel workbook
def findSemesters(ws):
    try:
        courses = 0
        # Iterate through cells in excel file
        for row, cells in enumerate(ws.values):
            for col, cell in enumerate(cells):
                # If cell matches a recognized semester format
                if cell != None and re.search('(fall|spring|summer) [0-9]{4}', str(cell), flags = re.IGNORECASE):
                    semester = re.search('(fall|spring|summer) [0-9]{4}', str(cell), flags = re.IGNORECASE).group(0)
                    # Check for courses directly underneath semester cell (up to limit)
                    for i in range(1, 99):
                        text = ws[get_column_letter(col + 1) + str(row + i)]
                        if text.value != None:
                            # Check that text matches required course format and then update data
                            for course in re.findall('[A-Z]{4} [0-9].{3,4}', str(text.value)):
                                data.addSemester(semester)
                                data.getSemester(semester).addCourse(course.rstrip())
                                courses += 1
                        else:
                            break
        return courses
    
    except IndexError:
        raise Exception("Unable to read cell, aborting.\nIs file corrupted?")
    except Exception:
        raise Exception("Unexpected error occurred reading from Excel, aborting.\nPlease check input files and try again.")



# Function to write to Excel file (using template)
def outputExcel(filename, excel, ws):
    try:
        semesters = data.getSortedSemesters()
        row, courses = 3, 3
        # While semesters are still being processed
        while len(semesters) > 0:
            for column in "ACE":
                # Force out of loop when no semesters remaining
                if semesters == []:
                    break
                # If semester not in correct column then continue
                if (column == "A" and "fall" not in semesters[0].season.lower()): continue
                if (column == "C" and "spring" not in semesters[0].season.lower()): continue
                if (column == "E" and "summer" not in semesters[0].season.lower()): break
                # Otherwise remove first semester from list
                semester = semesters.pop(0)
                # Update Excel file accordingly
                ws[column + str(row)] = f"{semester.season} {semester.year}"
                for index, course in enumerate(semester.courses):
                    ws[column + str(row + index + 1)].value = course + " - " + getCourse(course).desc
                    ws["H" + str(courses)].value = course + " - " + getCourse(course).desc
                    courses += 1
            row += 9
        # Save Excel file
        excel.save(filename)

    except IndexError:
        raise Exception("Unable to write cell, aborting.\nWas template file moved or deleted?")
    except Exception:
        raise Exception("Unexpected error occurred writing to file, aborting.\nIs template file corrupted?")